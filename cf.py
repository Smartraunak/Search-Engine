from bs4 import BeautifulSoup
import time
from openpyxl import Workbook
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

siteUrl = 'https://codeforces.com/problemset'
questionNamelist = []
questionUrlList = []
questionDifficultyList = []
qusetionDescription = []


def xcelSheet():

    excelFileName = 'Codeforces.xlsx'
    sheetName = 'Data Structures and Algorithms'
    df = pd.DataFrame({
        'Question Name': questionNamelist,
        'Question Url': questionUrlList,
        'Question Difficulty': questionDifficultyList,
        'Question Description': qusetionDescription
    })

    wb = Workbook()
    sheet1 = wb.create_sheet(sheetName)
    sheet1.cell(1, 1, 'Question Name')
    sheet1.cell(1, 2, 'Question Difficulty')
    sheet1.cell(1, 3, 'Question Description')
    sheet1.cell(1, 4, 'Question Url')

    for i in range(0, df.__len__()):
        sheet1.cell(i+2, 1, df['Question Name'][i])
        sheet1.cell(i+2, 2, df['Question Difficulty'][i])
        sheet1.cell(i+2, 3, df['Question Description'][i])
        sheet1.cell(i+2, 4, df['Question Url'][i])

    wb.save(excelFileName)
    wb.close()
    print(" Excel Sheet Created")


def openBrowser(url):
    print("---opening Browser---")
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.add_argument('--incognito')
    options.add_argument('--headless')

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),
                              options=options)

    driver.get(url)
    driver.maximize_window()
    return driver


def closeBrowser(driver):
    print(" ---- closing broswer")
    driver.close()


def fetchdes(pageUrl, title):
    sleepTime = 1
    browser = openBrowser(pageUrl)
    time.sleep(sleepTime)
    pageSource = browser.page_source
    WebDriverWait(browser, 10).until(EC.title_contains("Problem - "+title+" - Codeforces"))

    soup = BeautifulSoup(pageSource, 'html.parser')
    if (browser.title == "Problem - "+title+" - Codeforces"):
        print("----Parsing descriptiondata----")
        newSoup = BeautifulSoup(pageSource, 'html.parser')
        questiondis = newSoup.find('div', class_='problem-statement')
        question =questiondis.find_all('div',class_='')
        paragraphs = question[0].find_all('p',class_='')
        combined_text = ""
        for paragraph in paragraphs:
            text = paragraph.get_text()
            combined_text += text + " "
        combined_text=combined_text.strip()
        qusetionDescription.append(combined_text)
        print("------Done")
        closeBrowser(browser)
    else:
        print("Page does not exist. Connection failed, Status Code: ",
              soup.status_code)
    return


def fetchPageData(pageUrl):
    sleepTime = 2
    browser = openBrowser(pageUrl)
    time.sleep(sleepTime)
    pageSource = browser.page_source
    WebDriverWait(browser, 10).until(EC.title_contains("Problemset - Codeforces"))

    soup = BeautifulSoup(pageSource, 'html.parser')
    if (browser.title == "Problemset - Codeforces"):
        print("----Parsing pagedata----")
        newSoup = BeautifulSoup(pageSource, 'html.parser')
        questiontable = newSoup.find('table', class_='problems')
        questionBlock = questiontable.find('tbody')
        questionList = questionBlock.find_all('tr')

        first_question_encountered = False
        for question in questionList:
            row = question.find_all('td')
            if not first_question_encountered:
                first_question_encountered = True
                continue
            questiontitle=row[0].find('a').text.strip()
            questionUrl = row[0].find('a')['href']
            questionnam=row[1].find_all('div')
            questionName = questionnam[0].find('a').text
            questionUrl = 'https://codeforces.com' + questionUrl
            try:
               questionDifficulty = row[3].find('span').text
            except AttributeError:
                questionDifficulty="1500"
            fetchdes(questionUrl,questiontitle)
            questionNamelist.append(questionName)
            questionUrlList.append(questionUrl)
            questionDifficultyList.append(questionDifficulty)
        print("------Done")
        closeBrowser(browser)
    else:
        print("Page does not exist. Connection failed, Status Code: ",
              soup.status_code)
    return


def getData():

    try:
        browser = openBrowser(siteUrl)
        time.sleep(2)
        pageSource = browser.page_source
        WebDriverWait(browser, 10).until(
            EC.title_contains("Problemset - Codeforces"))
        soup = BeautifulSoup(pageSource, 'html.parser')

        if (browser.title == "Problemset - Codeforces"):
            totalPage = 85
            closeBrowser(browser)
            for page in range(1, totalPage+1):
                print("----fetching page-----")
                pageUrl = siteUrl + '/page/'+str(page)
                fetchPageData(pageUrl)

            print("---Done scrapping---")
            xcelSheet()
        else:
            print("Connection failed")
            return
    except Exception as e:
        print("Some error occured, error: ", e)
        return


if __name__ == "__main__":
    getData()
